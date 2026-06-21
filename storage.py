import csv
import json
import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class HourlyStorage:
    """Persists count events and aggregates into hour buckets for Excel export."""

    def __init__(self, db_path="counts.db"):
        self.db_path = db_path
        self._init_db()

    def _connect(self):
        return sqlite3.connect(self.db_path)

    def _init_db(self):
        with self._connect() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS hourly_counts (
                    hour_bucket TEXT PRIMARY KEY,
                    count INTEGER NOT NULL
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS events (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ts TEXT NOT NULL,
                    class TEXT NOT NULL DEFAULT 'object'
                )
            """)
            # Per-class hourly aggregation (Roadmap Phase 4).
            conn.execute("""
                CREATE TABLE IF NOT EXISTS class_hourly (
                    hour_bucket TEXT NOT NULL,
                    class TEXT NOT NULL,
                    count INTEGER NOT NULL,
                    PRIMARY KEY (hour_bucket, class)
                )
            """)
            # Best-effort migration: add the class column to a pre-existing
            # events table that predates per-class tracking.
            cols = [r[1] for r in conn.execute("PRAGMA table_info(events)")]
            if "class" not in cols:
                conn.execute(
                    "ALTER TABLE events ADD COLUMN class TEXT NOT NULL DEFAULT 'object'"
                )

    def increment(self, cls="object", n=1):
        cls = cls or "object"
        now = datetime.now()
        bucket = now.strftime("%Y-%m-%d %H:00")
        with self._connect() as conn:
            conn.execute(
                """INSERT INTO hourly_counts (hour_bucket, count) VALUES (?, ?)
                   ON CONFLICT(hour_bucket) DO UPDATE SET count = count + ?""",
                (bucket, n, n),
            )
            conn.execute(
                """INSERT INTO class_hourly (hour_bucket, class, count) VALUES (?, ?, ?)
                   ON CONFLICT(hour_bucket, class) DO UPDATE SET count = count + ?""",
                (bucket, cls, n, n),
            )
            conn.executemany(
                "INSERT INTO events (ts, class) VALUES (?, ?)",
                [(now.isoformat(), cls)] * n,
            )

    def increment_many(self, classes):
        """Record a batch of crossings given a list of class names."""
        for cls in classes:
            self.increment(cls)

    def get_current_hour_count(self):
        bucket = datetime.now().strftime("%Y-%m-%d %H:00")
        with self._connect() as conn:
            row = conn.execute(
                "SELECT count FROM hourly_counts WHERE hour_bucket = ?",
                (bucket,),
            ).fetchone()
        return row[0] if row else 0

    def get_total(self):
        with self._connect() as conn:
            row = conn.execute("SELECT SUM(count) FROM hourly_counts").fetchone()
        return row[0] or 0

    def get_all_hourly(self):
        with self._connect() as conn:
            return conn.execute(
                "SELECT hour_bucket, count FROM hourly_counts ORDER BY hour_bucket"
            ).fetchall()

    def get_class_breakdown(self):
        """All-time per-class totals as a list of (class, count), most first."""
        with self._connect() as conn:
            return conn.execute(
                """SELECT class, SUM(count) FROM class_hourly
                   GROUP BY class ORDER BY SUM(count) DESC"""
            ).fetchall()

    def get_current_hour_class_breakdown(self):
        bucket = datetime.now().strftime("%Y-%m-%d %H:00")
        with self._connect() as conn:
            return conn.execute(
                """SELECT class, count FROM class_hourly
                   WHERE hour_bucket = ? ORDER BY count DESC""",
                (bucket,),
            ).fetchall()

    def get_all_class_hourly(self):
        with self._connect() as conn:
            return conn.execute(
                """SELECT hour_bucket, class, count FROM class_hourly
                   ORDER BY hour_bucket, class"""
            ).fetchall()

    def get_daily(self):
        """All-time per-day totals as a list of (date, count)."""
        daily = {}
        for bucket, count in self.get_all_hourly():
            date_part = bucket.split(" ")[0]
            daily[date_part] = daily.get(date_part, 0) + count
        return sorted(daily.items())

    def reset(self):
        """Wipe all history (destructive)."""
        with self._connect() as conn:
            conn.execute("DELETE FROM hourly_counts")
            conn.execute("DELETE FROM class_hourly")
            conn.execute("DELETE FROM events")

    def clear_range(self, start_date, end_date):
        """Non-destructive clear: delete only buckets within [start_date, end_date]
        (inclusive, 'YYYY-MM-DD' strings). Returns the number of counts removed."""
        lo = f"{start_date} 00:00"
        hi = f"{end_date} 23:59"
        with self._connect() as conn:
            removed = conn.execute(
                "SELECT COALESCE(SUM(count), 0) FROM hourly_counts "
                "WHERE hour_bucket BETWEEN ? AND ?",
                (lo, hi),
            ).fetchone()[0]
            conn.execute(
                "DELETE FROM hourly_counts WHERE hour_bucket BETWEEN ? AND ?",
                (lo, hi),
            )
            conn.execute(
                "DELETE FROM class_hourly WHERE hour_bucket BETWEEN ? AND ?",
                (lo, hi),
            )
            conn.execute(
                "DELETE FROM events WHERE ts BETWEEN ? AND ?",
                (f"{start_date}T00:00:00", f"{end_date}T23:59:59.999999"),
            )
        return removed

    def export_excel(self, path):
        rows = self.get_all_hourly()
        wb = Workbook()

        ws = wb.active
        ws.title = "Hourly Counts"
        self._write_hourly_sheet(ws, rows)

        ws2 = wb.create_sheet("Daily Summary")
        self._write_daily_sheet(ws2, rows)

        ws3 = wb.create_sheet("Per-Class")
        self._write_class_sheet(ws3)

        wb.save(path)

    def export_csv(self, path):
        """Flat CSV: hourly buckets plus per-class breakdown rows."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["section", "date", "hour", "class", "count"])
            for bucket, count in self.get_all_hourly():
                date_part, hour_part = bucket.split(" ")
                writer.writerow(["hourly", date_part, hour_part, "", count])
            for bucket, cls, count in self.get_all_class_hourly():
                date_part, hour_part = bucket.split(" ")
                writer.writerow(["class_hourly", date_part, hour_part, cls, count])
            for cls, count in self.get_class_breakdown():
                writer.writerow(["class_total", "", "", cls, count])

    def export_json(self, path):
        """Structured JSON report: total, hourly, daily, and per-class."""
        report = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "total": self.get_total(),
            "hourly": [
                {"hour_bucket": b, "count": c} for b, c in self.get_all_hourly()
            ],
            "daily": [
                {"date": d, "count": c} for d, c in self.get_daily()
            ],
            "class_totals": [
                {"class": cls, "count": c} for cls, c in self.get_class_breakdown()
            ],
            "class_hourly": [
                {"hour_bucket": b, "class": cls, "count": c}
                for b, cls, c in self.get_all_class_hourly()
            ],
        }
        with open(path, "w") as f:
            json.dump(report, f, indent=2)

    @staticmethod
    def _header(cell):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2B7FFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_hourly_sheet(self, ws, rows):
        headers = ["Date", "Hour", "Count"]
        for col, h in enumerate(headers, 1):
            self._header(ws.cell(row=1, column=col, value=h))

        total = 0
        for i, (bucket, count) in enumerate(rows, 2):
            date_part, hour_part = bucket.split(" ")
            ws.cell(row=i, column=1, value=date_part)
            ws.cell(row=i, column=2, value=hour_part)
            ws.cell(row=i, column=3, value=count)
            total += count

        total_row = max(2, len(rows) + 2)
        tc1 = ws.cell(row=total_row, column=1, value="TOTAL")
        tc1.font = Font(bold=True)
        tc1.fill = PatternFill("solid", fgColor="EEF2F7")
        ws.cell(row=total_row, column=2).fill = PatternFill("solid", fgColor="EEF2F7")
        tc3 = ws.cell(row=total_row, column=3, value=total)
        tc3.font = Font(bold=True)
        tc3.fill = PatternFill("solid", fgColor="EEF2F7")

        for col, width in enumerate([14, 10, 12], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"

    def _write_class_sheet(self, ws):
        headers = ["Class", "Total Count"]
        for col, h in enumerate(headers, 1):
            self._header(ws.cell(row=1, column=col, value=h))

        breakdown = self.get_class_breakdown()
        total = 0
        for i, (cls, count) in enumerate(breakdown, 2):
            ws.cell(row=i, column=1, value=cls)
            ws.cell(row=i, column=2, value=count)
            total += count

        total_row = max(2, len(breakdown) + 2)
        tc1 = ws.cell(row=total_row, column=1, value="TOTAL")
        tc1.font = Font(bold=True)
        tc1.fill = PatternFill("solid", fgColor="EEF2F7")
        tc2 = ws.cell(row=total_row, column=2, value=total)
        tc2.font = Font(bold=True)
        tc2.fill = PatternFill("solid", fgColor="EEF2F7")

        for col, width in enumerate([20, 14], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"

    def _write_daily_sheet(self, ws, rows):
        headers = ["Date", "Total Count"]
        for col, h in enumerate(headers, 1):
            self._header(ws.cell(row=1, column=col, value=h))

        daily = {}
        for bucket, count in rows:
            date_part = bucket.split(" ")[0]
            daily[date_part] = daily.get(date_part, 0) + count

        for i, (date_part, count) in enumerate(sorted(daily.items()), 2):
            ws.cell(row=i, column=1, value=date_part)
            ws.cell(row=i, column=2, value=count)

        for col, width in enumerate([14, 14], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"
